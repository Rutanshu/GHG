"""
Parse India's Central Electricity Authority 'CO2 Baseline Database for the
Indian Power Sector' (Version 21.0) into normalized emission-factor records:
- National headline grid emission factors (Results sheet)
- Per-generating-unit specific CO2 emission rate, latest year (Data sheet)

Source: https://cea.nic.in/cdm-co2-baseline-database/?lang=en
File used: CO2_Database_V_21.0.xlsx
"""
import json
import re
import os
import openpyxl

SRC_PATH = "raw/cea-co2-database-v21.xlsx"
OUT_PATH = "out/cea-india-2025.json"

SOURCE_URL = "https://cea.nic.in/cdm-co2-baseline-database/?lang=en"
ORG = "Central Electricity Authority (CEA), Government of India"
DATASET = "CO2 Baseline Database for the Indian Power Sector, Version 21.0"
LICENCE = "Government of India — free to use with attribution"
PUB_YEAR = 2025

records = []
_seen_ids = set()


def add(rid, activity, value, unit_denom, region, source_page, year, notes=None):
    base = "cea-india-2025-" + re.sub(r"[^a-z0-9]+", "-", rid.lower()).strip("-")
    fid, n = base, 2
    while fid in _seen_ids:
        fid = f"{base}-{n}"
        n += 1
    _seen_ids.add(fid)
    records.append({
        "id": fid,
        "activity": activity,
        "scope": 2,
        "category": "2",
        "category_name": "Scope 2 purchased energy",
        "method": "activity-based",
        "value": round(float(value), 6),
        "unit_numerator": "tCO2",
        "unit_denominator": unit_denom,
        "gases": ["CO2"],
        "gwp_basis": "IPCC AR5 GWP100",
        "country": "IN",
        "region": region,
        "year": year,
        "publication_year": PUB_YEAR,
        "organization": ORG,
        "dataset": DATASET,
        "source_url": SOURCE_URL,
        "source_page_or_table": source_page,
        "licence": LICENCE,
        "price_year": None,
        "currency_deflator_note": None,
        "boundary": "generation (location-based grid average electricity)",
        "value_status": "verified",
        "notes": notes,
    })


def national_headline(wb):
    ws = wb["Results"]
    rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
    years_row = rows[11]  # row 12
    years = [y for y in years_row[5:8] if y]
    labels = {
        13: "Weighted Average Emission Rate",
        14: "Weighted Average Grid Emission Rate (incl. RES, Captive)",
        15: "Simple Operating Margin",
        16: "Build Margin",
        17: "Combined Margin",
    }
    for row_idx, label in labels.items():
        r = rows[row_idx - 1]
        for i, year in enumerate(years):
            val = r[5 + i]
            if val is None:
                continue
            add(f"national-{label}-{year}", f"India national grid — {label}", val, "MWh",
                "India (national)", f"Results sheet — Emission Factors table, row '{label}', FY {year}",
                int(str(year).split("-")[0]),
                notes=f"Financial year {year}. Methodology: {DATASET.split(', ')[0]} per ACM0002 / UNFCCC Tool 07.")
    n = len(records)
    print(f"CEA national headline: {n} records")


def per_unit(wb):
    ws = wb["Data"]
    rows = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows)
    # Locate columns by fuzzy header match (headers contain newlines)
    def col(*keywords):
        for i, h in enumerate(header):
            if h and all(k.lower() in str(h).lower() for k in keywords):
                return i
        return None

    name_c, state_c, sector_c, fuel_c = 1, 5, 6, 9
    gen_c = col("Net", "Generation", "GWh")
    spec_c = col("Specific", "Emissions")
    year_label = None
    for h in header:
        if h and "Specific" in str(h):
            m = re.search(r"(\d{4}-\d{2})", str(h))
            if m:
                year_label = m.group(1)
    n = 0
    for r in rows:
        if r[0] is None:
            continue
        name, state, sector, fuel = r[name_c], r[state_c], r[sector_c], r[fuel_c]
        gen, spec = r[gen_c], r[spec_c]
        if spec in (None, 0) or gen in (None, 0):
            continue
        activity = f"Purchased electricity — {name} ({state}, {sector}, {fuel})"
        add(f"unit-{r[0]}-{name}-{r[2]}", activity, spec, "MWh", state,
            f"Data sheet — {name}, unit {r[2]}, FY {year_label} specific emissions",
            int(year_label.split("-")[0]) if year_label else 2024,
            notes=f"Generating unit type: {sector}. Primary fuel: {fuel}. FY {year_label} net generation {gen:.1f} GWh.")
        n += 1
    print(f"CEA per-unit: {n} records")


def main():
    wb = openpyxl.load_workbook(SRC_PATH, read_only=True, data_only=True)
    national_headline(wb)
    per_unit(wb)

    print(f"CEA total: {len(records)}")
    ids = [r["id"] for r in records]
    assert len(ids) == len(set(ids)), "duplicate ids in CEA output!"

    os.makedirs("out", exist_ok=True)
    with open(OUT_PATH, "w") as f:
        json.dump(records, f, ensure_ascii=False)
    print("wrote", OUT_PATH)


if __name__ == "__main__":
    main()
